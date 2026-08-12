"""Reporting Agent (Agent 3) - the consolidated, ring-grouped Excel report.

No LLM call anywhere in this file. Every column the report needs (serial
number, transaction id, token, real account number, amount, credited-from,
sent-to, timestamp, day of week, ML score) is already a known fact sitting
in Neo4j, Redis, or the known customer/ground-truth population - none of it
is generated or judged, so an LLM has nothing useful to add to the table
itself. Same principle as the Money-Trail Agent's own core: never trust an
LLM to retype a fact deterministic code can compute reliably.

Fires the moment agents/money_trail_agent.py confirms convergence_found -
independent of the Verdict Agent's own GUILTY/NOT_GUILTY opinion, which
separately and exclusively decides FL label-writing (see CLAUDE.md's
"Post-Session 6 Extension - Reframed 3-Agent Pipeline"). A structurally
confirmed ring is always worth a human's five minutes to look at; a
silently-dropped one is not recoverable.

One real, honest limitation for this demo project: "respective account
number" needs the real, unmasked account number, which nothing upstream of
this file ever touches (every other component only ever sees token_id, by
design - see CLAUDE.md's masking decision). This file is deliberately the
first and only place in the whole pipeline allowed to cross back over that
boundary - reversing token_id -> real account number here only works because
this is a synthetic demo with the full account population sitting in known
data files (customers.json + ground-truth files), so a lookup can be built
by re-hashing every known account number and matching. A real production
system would resolve this from each branch's own local account_master
instead, never from a shared cross-branch table.
"""
import json
import os
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

from branch_node.masking import token_for
from graph.queries import get_incoming_txns, get_outgoing_txns
from shared.config import CASH_SINK_TOKEN, CASH_SOURCE_TOKEN

_SENTINELS = {CASH_SOURCE_TOKEN, CASH_SINK_TOKEN, "MERCHANT"}

_CUSTOMER_FILES = [
    "data/customers.json",
    "data/scenario_500/customers.json",
    "data/scenario_1500/customers.json",
]
_GROUND_TRUTH_FILES = [
    "data/layering_hops2_ground_truth.json",
    "data/layering_hops4_ground_truth.json",
    "data/layering_hops6_ground_truth.json",
    "data/scenario_500/multi_ring_ground_truth.json",
    "data/scenario_1500/multi_ring_ground_truth.json",
]

_TXN_HEADERS = [
    "Ring ID", "Serial No", "Transaction ID", "Token", "Account Number", "Amount",
    "Credited From", "Sent To", "Time of Transaction", "Day of Week", "ML Model Score",
    "Verdict", "Verdict Confidence", "Caught At",
]
_SUMMARY_HEADERS = [
    "Ring ID", "Started By", "All Tokens Involved", "Converged To",
    "Num Sources", "Num Branches", "Amount Preservation Ratio", "Total Transactions",
    "Verdict", "Verdict Confidence", "Verdict Rationale", "Caught At",
]


def _parse_ts(ts: str):
    try:
        return datetime.strptime(ts, "%Y-%m-%dT%H:%M:%S.%fZ")
    except (ValueError, TypeError):
        return None


def _build_account_lookup() -> dict:
    """token_id -> real account_number, built by re-hashing every account
    number in the known synthetic population - see module docstring for why
    this only works because this is a demo with the full population on disk.
    """
    lookup = {}
    for path in _CUSTOMER_FILES:
        if not os.path.exists(path):
            continue
        data = json.load(open(path))
        for acct in data.get("accounts", []):
            num = acct["account_number"]
            lookup[token_for(num)] = num
    for path in _GROUND_TRUTH_FILES:
        if not os.path.exists(path):
            continue
        gt = json.load(open(path))
        accounts = gt.get("fraud_accounts") or gt.get("all_fraud_accounts") or []
        for num in accounts:
            lookup[token_for(num)] = num
    return lookup


def _resolve_account(lookup: dict, token: str) -> str:
    if token in _SENTINELS:
        return token
    return lookup.get(token, f"UNKNOWN({token[:10]})")


def _fetch_score_record(r, txn_id: str):
    """The real ScoreRecord this exact transaction was scored under - gives
    the authoritative token_id for this row (which side of the transfer was
    "local" per branch_node/masking.py) rather than guessing.
    """
    keys = r.keys(f"score:*:*:{txn_id}")
    if not keys:
        return None
    raw = r.get(keys[0])
    return json.loads(raw) if raw else None


def _gather_ring_transactions(driver, convergence_result: dict) -> list:
    """Walks Neo4j directly (fresh reads, not reused from anywhere else) to
    build the COMPLETE transaction list for this ring: every source's
    original CASH deposit, every intermediate hop already in
    convergence_result['paths'], and the final cash-out from the
    convergence account - none of which money_trail_agent.py's own
    evidence['path'] includes on its own (it only covers the hops between
    sources and the convergence account, not the deposit or exit legs).
    """
    rows = []
    seen_txn_ids = set()
    convergence_node = convergence_result["convergence_account"]

    def _add(from_token, to_token, edge):
        if edge["txn_id"] in seen_txn_ids:
            return
        seen_txn_ids.add(edge["txn_id"])
        rows.append({
            "from_token": from_token,
            "to_token": to_token,
            "txn_id": edge["txn_id"],
            "amount": edge["amount"],
            "ts": edge["ts"],
        })

    for path in convergence_result["paths"]:
        source_token = path[0]
        cash_edges = [e for e in get_incoming_txns(driver, source_token) if e["from_token"] == CASH_SOURCE_TOKEN]
        if cash_edges:
            _add(CASH_SOURCE_TOKEN, source_token, cash_edges[0])
        for from_tok, to_tok in zip(path, path[1:]):
            match = next((e for e in get_outgoing_txns(driver, from_tok) if e["to_token"] == to_tok), None)
            if match:
                _add(from_tok, to_tok, match)

    cash_out_edges = [e for e in get_outgoing_txns(driver, convergence_node) if e["to_token"] == CASH_SINK_TOKEN]
    if cash_out_edges:
        _add(convergence_node, CASH_SINK_TOKEN, cash_out_edges[0])

    rows.sort(key=lambda row: row["ts"])
    return rows


def _load_or_create_workbook(path: str):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(_TXN_HEADERS)
    summary_ws = wb.create_sheet("Ring Summaries")
    summary_ws.append(_SUMMARY_HEADERS)
    return wb


def _refresh_ring_timestamp(path: str, ring_id: str, caught_at: datetime) -> bool:
    """If this ring already has rows in the report, bump their "Caught At"
    cell to now instead of silently doing nothing on a repeat run - the old
    behavior (skip entirely once a ring_id is seen once) looked indistinguishable
    from a broken report to anyone re-running the same scenario later in the
    day. Returns True if the ring was found and updated (so the caller can
    skip re-deriving the transaction data from Neo4j/Redis entirely - the
    transactions themselves haven't changed, only "when did we last see
    this"), False if this ring isn't in the report yet and needs full rows
    written for the first time.
    """
    if not os.path.exists(path):
        return False
    wb = openpyxl.load_workbook(path)
    if "Ring Summaries" not in wb.sheetnames:
        return False

    summary_ws = wb["Ring Summaries"]
    summary_caught_col = summary_ws.max_column  # "Caught At" is always the last column
    found = False
    for row in summary_ws.iter_rows(min_row=2):
        if row[0].value == ring_id:
            row[summary_caught_col - 1].value = caught_at
            found = True
            break
    if not found:
        return False

    txn_ws = wb["Transactions"]
    txn_caught_col = txn_ws.max_column
    for row in txn_ws.iter_rows(min_row=2):
        if row[0].value == ring_id:
            row[txn_caught_col - 1].value = caught_at

    wb.save(path)
    return True


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 50)


def _write_excel(path: str, ring_id: str, summary_row: dict, txn_rows: list, caught_at: datetime) -> None:
    wb = _load_or_create_workbook(path)

    txn_ws = wb["Transactions"]
    for serial, row in enumerate(txn_rows, start=1):
        txn_ws.append([
            ring_id, serial, row["txn_id"], row["token"], row["account_number"],
            row["amount"], row["credited_from"], row["sent_to"], row["time"],
            row["day_of_week"], row["ml_score"], summary_row["verdict"], summary_row["verdict_confidence"],
            caught_at,
        ])

    summary_ws = wb["Ring Summaries"]
    summary_ws.append([
        summary_row["ring_id"], summary_row["started_by"], summary_row["all_tokens"],
        summary_row["converged_to"], summary_row["num_sources"], summary_row["num_branches"],
        summary_row["amount_preservation_ratio"], summary_row["total_transactions"],
        summary_row["verdict"], summary_row["verdict_confidence"], summary_row["verdict_rationale"],
        caught_at,
    ])

    _autosize_columns(txn_ws)
    _autosize_columns(summary_ws)
    wb.save(path)


def report_agent(driver, r, convergence_result: dict, ring_id: str, verdict: dict = None) -> str:
    """Builds/appends this confirmed ring's full transaction list to the
    single consolidated Excel report (default reports/fraud_rings_report.xlsx,
    override with REPORT_XLSX_PATH). Never duplicates a ring's rows - but a
    repeat run (e.g. re-streaming the same scenario later in the day) DOES
    still do something visible: it bumps that ring's "Caught At" cell to the
    current system time (real wall-clock time, not the transaction's own
    fabricated timestamp) rather than silently no-op'ing, which used to look
    indistinguishable from the report being broken.

    verdict (optional, {verdict, confidence, rationale}) is displayed only -
    it never decides whether this ring gets reported (this function already
    only ever runs on a confirmed convergence_found; see money_trail_agent.py
    for why reporting is deliberately independent of the Verdict Agent's
    opinion). Pass None (e.g. from the standalone __main__ smoke test below,
    which has no verdict on hand) and the report just shows "N/A".
    """
    output_path = os.getenv("REPORT_XLSX_PATH", "reports/fraud_rings_report.xlsx")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    caught_at = datetime.now()

    if _refresh_ring_timestamp(output_path, ring_id, caught_at):
        return output_path

    lookup = _build_account_lookup()
    hops = _gather_ring_transactions(driver, convergence_result)

    txn_rows = []
    for hop in hops:
        record = _fetch_score_record(r, hop["txn_id"])
        # Fallback when the score record is missing (shouldn't normally happen,
        # but never let it silently write a sentinel like CASH_OUT as if it
        # were a real account): prefer whichever side of the hop isn't a
        # sentinel, instead of always defaulting to to_token.
        fallback_token = hop["from_token"] if hop["to_token"] in _SENTINELS else hop["to_token"]
        token = record["token_id"] if record else fallback_token
        score = record["score"] if record else None
        ts = _parse_ts(hop["ts"])
        txn_rows.append({
            "txn_id": hop["txn_id"],
            "token": token,
            "account_number": _resolve_account(lookup, token),
            "amount": hop["amount"],
            "credited_from": _resolve_account(lookup, hop["from_token"]),
            "sent_to": _resolve_account(lookup, hop["to_token"]),
            "time": ts,
            "day_of_week": ts.strftime("%A") if ts else "",
            "ml_score": score,
        })

    source_tokens = [path[0] for path in convergence_result["paths"]]
    all_tokens = sorted({tok for path in convergence_result["paths"] for tok in path})
    verdict = verdict or {}
    summary_row = {
        "ring_id": ring_id,
        "started_by": ", ".join(_resolve_account(lookup, t) for t in source_tokens),
        "all_tokens": ", ".join(all_tokens),
        "converged_to": _resolve_account(lookup, convergence_result["convergence_account"]),
        "num_sources": convergence_result.get("num_sources"),
        "num_branches": convergence_result.get("num_branches"),
        "amount_preservation_ratio": convergence_result.get("amount_preservation_ratio"),
        "total_transactions": len(txn_rows),
        "verdict": verdict.get("verdict", "N/A"),
        "verdict_confidence": verdict.get("confidence"),
        "verdict_rationale": verdict.get("rationale", ""),
    }

    _write_excel(output_path, ring_id, summary_row, txn_rows, caught_at)
    return output_path


if __name__ == "__main__":
    import sys

    import redis

    from graph.connection import get_driver
    from shared.config import REDIS_URL

    from agents.money_trail_agent import _find_convergence_group_for_token

    token = sys.argv[1] if len(sys.argv) > 1 else None
    if not token:
        raise SystemExit("Usage: python -m agents.report_agent <token_id>  (must resolve to a real convergence)")

    driver = get_driver()
    r = redis.from_url(REDIS_URL)
    result = _find_convergence_group_for_token(driver, r, token)
    if not result.get("has_convergence"):
        raise SystemExit(f"No confirmed convergence for {token}: {result}")

    ring_id = f"ring_{result['convergence_account'][:8]}"
    path = report_agent(driver, r, result, ring_id)
    print(f"Report written -> {path}")
    driver.close()
