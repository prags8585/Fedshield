"""Shared agent-orchestration flow, reused by Tab 2 (216 case) and Tab 3
(1500 case) -- the only thing that differs between them is which producer
command gets fired. The investigation feed (Redis evidence/verdicts/
flagged_accounts/structuring_log) is still scenario-agnostic and shared
between the two tabs -- the listener and agents don't know or care which
scenario a flagged token came from. What is NOT shared any more is each
tab's own terminals: Tab 2 and Tab 3 each get their own master/producer/
branch-log PTY sessions (tab2-* / tab3-* channel names, in their own
`agent_registry`, served over their own /ws/agents/{channel} route below)
so running one tab's terminals never shows up in the other's.
"""
import asyncio
import json
import os
from pathlib import Path

import openpyxl
import redis
from fastapi import APIRouter, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse

from shared.redis_keys import FLAGGED_ACCOUNTS, STRUCTURING_LOG, verdict_key

from . import tab1
from .pty_manager import PtyRegistry

REPO_ROOT = Path(__file__).resolve().parents[2]

# One independent set of 5 channels per tab (master/producer/branch-loc1-3),
# in their own registry -- deliberately NOT tab1.registry, so Tab 1's own
# Reset (which calls tab1.registry.restart_all()) can never touch these, and
# so a Tab 2 Reset/Start can never touch Tab 3's sessions or vice versa.
_AGENT_TABS = ("tab2", "tab3")
_AGENT_CHANNEL_SUFFIXES = ("master", "producer", "branch-loc1", "branch-loc2", "branch-loc3")


def _channels_for(tab_id: str) -> dict:
    return {suffix: f"{tab_id}-{suffix}" for suffix in _AGENT_CHANNEL_SUFFIXES}


_ALL_AGENT_CHANNEL_NAMES = [
    f"{tab_id}-{suffix}" for tab_id in _AGENT_TABS for suffix in _AGENT_CHANNEL_SUFFIXES
]

agent_registry = PtyRegistry(
    _ALL_AGENT_CHANNEL_NAMES, cwd=str(REPO_ROOT), env=tab1._env, init_commands=tab1._init_commands,
)

ws_router = APIRouter()


@ws_router.websocket("/ws/agents/{channel_name}")
async def ws_agent_channel(websocket: WebSocket, channel_name: str):
    if channel_name not in agent_registry.sessions:
        await websocket.close(code=4404)
        return
    await websocket.accept()
    session = agent_registry.get(channel_name)
    if session.master_fd is None:
        session.start()
    if session.scrollback:
        await websocket.send_bytes(bytes(session.scrollback))
    session.subscribers.add(websocket)
    try:
        while True:
            message = await websocket.receive()
            if message["type"] == "websocket.disconnect":
                break
            data = message.get("bytes")
            if data is not None:
                session.write(data)
                continue
            text = message.get("text")
            if text is None:
                continue
            try:
                control = json.loads(text)
            except json.JSONDecodeError:
                continue
            if control.get("type") == "resize":
                session.resize(int(control.get("rows", 24)), int(control.get("cols", 80)))
    except WebSocketDisconnect:
        pass
    finally:
        session.subscribers.discard(websocket)

_redis = redis.Redis(host="localhost", port=6380, decode_responses=True)

# Same default agents/report_agent.py itself falls back to when REPORT_XLSX_PATH
# isn't set -- one shared workbook across both tabs, same as Redis's evidence/
# verdict keys are already shared, not scoped per scenario.
_DEFAULT_REPORT_PATH = REPO_ROOT / "reports" / "fraud_rings_report.xlsx"

LISTENER_CMD = (
    "NEO4J_URI=bolt://localhost:7688 REDIS_URL=redis://localhost:6380 "
    "PYTHONPATH=. python3 -u orchestrator/listener.py\n"
).encode()

BRANCH_LOG_CHANNELS = {
    "branch-loc1": "fedshieldv2-branch-loc1",
    "branch-loc2": "fedshieldv2-branch-loc2",
    "branch-loc3": "fedshieldv2-branch-loc3",
}

_DOCKER_SERVICES = {"redis", "neo4j", "kafka", "branch_loc1", "branch_loc2", "branch_loc3"}


async def docker_healthy() -> bool:
    try:
        proc = await asyncio.create_subprocess_exec(
            "docker", "compose", "ps", "--format", "json",
            cwd=str(REPO_ROOT),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )
        out, _ = await proc.communicate()
    except FileNotFoundError:
        return False

    text = out.decode(errors="replace").strip()
    if not text:
        return False

    rows = []
    try:
        parsed = json.loads(text)
        rows = parsed if isinstance(parsed, list) else [parsed]
    except json.JSONDecodeError:
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue

    healthy = set()
    for row in rows:
        name = row.get("Service")
        if name not in _DOCKER_SERVICES:
            continue
        health = row.get("Health", "")
        state = row.get("State", "")
        if health == "healthy" or (not health and state == "running"):
            healthy.add(name)
    return healthy.issuperset(_DOCKER_SERVICES)


async def _wait_for_subscribed(session, timeout: float = 15.0) -> bool:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + timeout
    while loop.time() < deadline:
        if b"subscribed to" in session.scrollback:
            return True
        await asyncio.sleep(0.3)
    return False


async def start_agents(producer_cmd: bytes, tab_id: str) -> dict:
    if not await docker_healthy():
        raise HTTPException(
            409,
            "docker stack not healthy -- complete Tab 1's data-gen/up sequence first",
        )

    channels = _channels_for(tab_id)

    master = agent_registry.get(channels["master"])
    if master.master_fd is None:
        master.start()

    # A bare shell already exists the moment this channel's WebSocket first
    # connects (same as Tab 1's terminals) -- that's not the same as the
    # listener actually running, so check scrollback content, not just
    # whether the PTY process exists.
    if b"subscribed to" not in master.scrollback:
        master.write(LISTENER_CMD)
        confirmed = await _wait_for_subscribed(master)
        if not confirmed:
            raise HTTPException(
                500,
                "listener did not confirm subscription within 15s -- check the master terminal",
            )

    # Start each branch's `docker logs -f` before the producer, if it isn't
    # already tailing -- otherwise this tab's own branch terminals sit blank
    # even though the containers are actively scoring, since a terminal only
    # ever shows something once a `docker logs -f` has actually been typed/
    # sent into it (same manual-shell design as Tab 1).
    for suffix, container in BRANCH_LOG_CHANNELS.items():
        branch = agent_registry.get(channels[suffix])
        if branch.master_fd is None:
            branch.start()
        cmd = f"docker logs -f {container}\n".encode()
        if cmd.strip() not in branch.scrollback:
            branch.write(cmd)

    producer = agent_registry.get(channels["producer"])
    if producer.master_fd is None:
        producer.start()
    producer.write(producer_cmd)

    return {"status": "started"}


async def reset_agents(tab_id: str) -> dict:
    # Restart only THIS tab's own 5 terminals -- never the other tab's, and
    # never Tab 1's/Tab 4's (those live in a different registry entirely).
    channels = _channels_for(tab_id)
    for suffix in _AGENT_CHANNEL_SUFFIXES:
        agent_registry.get(channels[suffix]).start()

    # The investigation feed itself (evidence/verdicts/flagged_accounts/
    # structuring_log/labels) is still shared across both tabs, unscoped --
    # this part is unchanged from before Reset was split by tab, and affects
    # whichever tab's Reset was clicked exactly as it always has.
    _redis.delete(FLAGGED_ACCOUNTS, STRUCTURING_LOG)
    for pattern in ("evidence:*", "verdicts:*", "labels:*"):
        keys = list(_redis.scan_iter(pattern))
        if keys:
            _redis.delete(*keys)
    return {"status": "reset"}


async def get_investigations() -> dict:
    items = []
    for key in _redis.scan_iter("evidence:*"):
        raw = _redis.get(key)
        if not raw:
            continue
        try:
            evidence = json.loads(raw)
        except json.JSONDecodeError:
            continue

        token_id = evidence.get("token_id") or key.split(":", 1)[1]
        structuring = None
        if evidence.get("structuring_reasoning") is not None:
            structuring = {
                "reasoning": evidence.get("structuring_reasoning"),
                "confidence": evidence.get("structuring_confidence"),
            }

        # money_trail_agent.py only embeds "verdict" into the evidence dict
        # for the token whose own investigation triggered the convergence --
        # every other group member gets the same evidence (stop_reason,
        # summary) but not that embedded copy. It does separately write
        # verdicts:{token_id} for every group member though, so fall back to
        # that key to give siblings their verdict too.
        verdict = evidence.get("verdict")
        if verdict is None:
            verdict_raw = _redis.get(verdict_key(token_id))
            if verdict_raw:
                try:
                    verdict = json.loads(verdict_raw)
                except json.JSONDecodeError:
                    verdict = None

        items.append({
            "token_id": token_id,
            "investigated_at": evidence.get("investigated_at"),
            "structuring": structuring,
            "money_trail": {
                "stop_reason": evidence.get("stop_reason"),
                "summary": evidence.get("summary"),
            },
            "verdict": verdict,
        })

    items.sort(key=lambda it: it["investigated_at"] or "", reverse=True)
    return {"items": items}


async def get_structuring_log(limit: int = 300) -> dict:
    """Agent 1's whole output now: a running list, not a per-token verdict --
    see CLAUDE.md's "Post-Session 6 Extension - Reframed 3-Agent Pipeline".
    Every flagged transaction (real fraud and false positive alike) gets one
    entry here, in arrival order, regardless of what Money-Trail later
    decides. Returns the most recent `limit` entries, newest first.
    """
    raw_items = _redis.lrange(STRUCTURING_LOG, -limit, -1)
    items = []
    for raw in raw_items:
        try:
            items.append(json.loads(raw))
        except json.JSONDecodeError:
            continue
    items.reverse()
    return {"items": items, "total": _redis.llen(STRUCTURING_LOG)}


async def get_rings() -> dict:
    """Money-Trail Agent's output, grouped for display the way the pipeline
    actually resolves it: every member of a confirmed ring gets the identical
    evidence written to its own evidence:{token_id} key (see
    money_trail_agent.py's group-write step), so scanning evidence:* and
    showing one card per token would just repeat the same ring N times.
    Groups by ring_id instead -- one card per real ring -- and separately
    lists tokens that never converged (insufficient_evidence/cycle/etc).

    Only the ONE token whose own investigation triggered the convergence
    carries "verdict"/"report_path" in its own evidence write (the rest are
    written from the pre-verdict evidence dict, before those fields exist on
    it) -- aggregating across every member of the group here recovers them
    without needing a per-token fallback.
    """
    rings_by_id: dict = {}
    dead_ends = []

    for key in _redis.scan_iter("evidence:*"):
        raw = _redis.get(key)
        if not raw:
            continue
        try:
            evidence = json.loads(raw)
        except json.JSONDecodeError:
            continue

        token_id = evidence.get("token_id") or key.split(":", 1)[1]

        if evidence.get("stop_reason") == "convergence_found":
            ring_id = evidence.get("ring_id") or f"ring_{evidence.get('convergence_node', 'unknown')}"
            ring = rings_by_id.setdefault(ring_id, {
                "ring_id": ring_id,
                "convergence_node": None,
                "source_tokens": None,
                "all_tokens": None,
                "summary": None,
                "verdict": None,
                "report_path": None,
                "investigated_at": None,
            })
            ring["convergence_node"] = ring["convergence_node"] or evidence.get("convergence_node")
            ring["source_tokens"] = ring["source_tokens"] or evidence.get("source_tokens")
            ring["all_tokens"] = ring["all_tokens"] or evidence.get("all_tokens")
            ring["summary"] = ring["summary"] or evidence.get("summary")
            if evidence.get("verdict") is not None:
                ring["verdict"] = evidence.get("verdict")
            if evidence.get("report_path") is not None:
                ring["report_path"] = evidence.get("report_path")
            investigated_at = evidence.get("investigated_at")
            if investigated_at and (ring["investigated_at"] is None or investigated_at > ring["investigated_at"]):
                ring["investigated_at"] = investigated_at
        else:
            dead_ends.append({
                "token_id": token_id,
                "stop_reason": evidence.get("stop_reason"),
                "summary": evidence.get("summary"),
                "investigated_at": evidence.get("investigated_at"),
            })

    # Defensive fallback only -- verdict_key is written for every group member
    # (money_trail_agent.py), so this should never actually be needed, but a
    # ring resolved mid-write (investigation still in flight) could otherwise
    # show a blank verdict for a moment.
    for ring in rings_by_id.values():
        if ring["verdict"] is None and ring["all_tokens"]:
            for tok in ring["all_tokens"]:
                vraw = _redis.get(verdict_key(tok))
                if vraw:
                    try:
                        ring["verdict"] = json.loads(vraw)
                        break
                    except json.JSONDecodeError:
                        continue

    rings = sorted(rings_by_id.values(), key=lambda r: r["investigated_at"] or "", reverse=True)
    dead_ends.sort(key=lambda d: d["investigated_at"] or "", reverse=True)
    return {"rings": rings, "dead_ends": dead_ends}


def _report_xlsx_path() -> Path:
    return Path(os.environ.get("REPORT_XLSX_PATH", str(_DEFAULT_REPORT_PATH)))


async def get_report_status() -> dict:
    path = _report_xlsx_path()
    if not path.exists():
        return {"ready": False}

    ring_count = 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        if "Ring Summaries" in wb.sheetnames:
            ring_count = sum(1 for _ in wb["Ring Summaries"].iter_rows(min_row=2, values_only=True))
    except Exception:
        pass

    return {"ready": True, "ring_count": ring_count, "mtime": path.stat().st_mtime}


async def download_report() -> FileResponse:
    path = _report_xlsx_path()
    if not path.exists():
        raise HTTPException(404, "report not generated yet -- no ring has converged")
    return FileResponse(
        str(path),
        filename="fraud_rings_report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
